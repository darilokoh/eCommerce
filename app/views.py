from django.shortcuts import render, redirect
from .forms import CambiarPasswordForm, ContactForm, ProductForm, CategoryForm, QueryTypeForm, RentalOrderForm, RecuperarForm, OrderForm
from django.contrib import messages
from datetime import timedelta
from django.contrib.auth import authenticate, login
from .models import Product, Category, Contact, QueryType, RentalOrder, RentalOrderItem, Order, OrderItem, Tokens, Region, Municipality
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import Http404, HttpResponse, JsonResponse, HttpRequest
from rest_framework import viewsets, serializers
from .serializers import ProductSerializer, CategorySerializer, ContactSerializer, QueryTypeSerializer, RentalOrderSerializer, RentalOrderItemSerializer, LoginSerializer, RegionSerializer, MunicipalitySerializer
import requests
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from app.cart import Cart
from rest_framework.response import Response
from django.conf import settings
from django.db.models import Sum, Q
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.decorators import api_view
from django.middleware.csrf import get_token
from django.views.decorators.http import require_http_methods
from .serializers import TokenSerializer
from .forms import UsuariosForm, LoginForm
from django.contrib.auth.models import User
import requests
from rest_framework.authtoken.models import Token
from rest_framework.views import APIView
from django.utils import timezone
from dateutil.parser import parse
from collections import Counter
from django.contrib.auth.hashers import make_password

# IMPORTS API TRANSBANK
from transbank.webpay.webpay_plus.transaction import Transaction, WebpayOptions
from transbank.common.integration_type import IntegrationType

# API HELPERS
from .api_helpers import LocationAPI, ProductAPI

from django.utils.crypto import get_random_string
tok = None


def is_staff(user):
    return (user.is_authenticated and user.is_superuser)

def is_admin(user):
    return user.groups.filter(name='admin').exists()

# VIEWSETS PARA APIS


class CategoryViewset(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer


class ProductViewset(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

    def get_queryset(self):
        products = Product.objects.all()

        name = self.request.GET.get('name')
        is_featured = self.request.GET.get('is_featured')
        category = self.request.GET.get('category')
        is_new = self.request.GET.get('is_new')
        min_price = self.request.GET.get('min_price_filter')
        max_price = self.request.GET.get('max_price_filter')
        is_rentable = self.request.GET.get('is_rentable')

        if name:
            products = products.filter(name__contains=name)
        if category:
            products = products.filter(category=category)
        if min_price and max_price:
            products = products.filter(price__range=(min_price, max_price))
        elif min_price:
            products = products.filter(price__gte=min_price)
        elif max_price:
            products = products.filter(price__lte=max_price)

        # Aplicar los filtros de featured y new
        if is_featured:
            products = products.filter(is_featured=True)
        if is_new:
            products = products.filter(is_new=True)
        # filtro para rentable
        if is_rentable:
            products = products.filter(is_rentable=True)

        return products

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=201, headers=headers)

    def perform_create(self, serializer):
        serializer.save()
    from django.contrib.auth import update_session_auth_hash

@require_http_methods(["GET", "POST"])
@login_required
def CambiarPassword(request):
    if request.method == 'POST':
        form = CambiarPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()

            messages.success(
                request, 'Tu contraseña ha sido actualizada exitosamente.')
            # Redirige a la página de inicio o donde prefieras
            return redirect('home')
        else:
            messages.error(
                request, 'Por favor corrige los errores a continuación.')
    else:
        form = CambiarPasswordForm(user=request.user)

    return render(request, 'registration/CambiarPassword.html', {'form': form})

class ContactViewSet(viewsets.ModelViewSet):
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer

    def get_queryset(self):
        queryset = Contact.objects.all()

        # Obtener el parámetro de consulta 'status' de la URL
        status = self.request.query_params.get('status', None)
        if status is not None and status != '':
            # Filtrar los contactos por estado
            queryset = queryset.filter(status=status)

        # Obtener el parámetro de consulta 'query_type_id' de la URL
        query_type_id = self.request.query_params.get('query_type_id', None)
        if query_type_id is not None and query_type_id != '':
            # Filtrar los contactos por ID del tipo de consulta
            queryset = queryset.filter(query_type_id=query_type_id)

        return queryset

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        status = request.data.get('status')
        instance.status = status
        instance.save()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

class QueryTypeViewset(viewsets.ModelViewSet):
    queryset = QueryType.objects.all()
    serializer_class = QueryTypeSerializer


class RentalOrderViewSet(viewsets.ModelViewSet):
    # def list(self, request):
    #     rental_orders = RentalOrder.objects.all()
    #     serializer = RentalOrderSerializer(rental_orders, many=True)
    #     return Response(serializer.data)
    queryset = RentalOrder.objects.all()
    serializer_class = RentalOrderSerializer


class RentalOrderItemViewSet(viewsets.ModelViewSet):
    queryset = RentalOrderItem.objects.all()
    serializer_class = RentalOrderItemSerializer

class RegionViewSet(viewsets.ReadOnlyModelViewSet):  # Solo lectura
    queryset = Region.objects.all()
    serializer_class = RegionSerializer

class MunicipalityViewSet(viewsets.ReadOnlyModelViewSet):  # Solo lectura
    serializer_class = MunicipalitySerializer

    def get_queryset(self):
        region_id = self.request.query_params.get('region_id')
        if region_id:
            return Municipality.objects.filter(region_id=region_id)
        return Municipality.objects.all()

# VISTAS INICIALES
@require_http_methods(["GET"])
def home(request):
    # Definimos los parámetros para filtrar productos
    params = {
        'is_new__in': 'true,false',
        'is_featured__in': 'true,false',
    }
    # Obtenemos los productos desde la API aplicando los filtros
    product_response = requests.get(
        settings.API_BASE_URL + 'product/', params=params).json()
    # Filtrar los productos para excluir los que tienen is_rentable=True
    filtered_products = [
        product for product in product_response if not product['is_rentable']]

    categories_response = requests.get(
        settings.API_BASE_URL + 'category/').json()

    data = {
        'products': filtered_products,
        'categories': categories_response
    }

    return render(request, 'app/home.html', data)

@require_http_methods(["GET", "POST"])
@api_view(['GET', 'POST'])
def catalogue(request):
    # Obtenemos los filtros desde el html
    name_filter = request.GET.get('name', '')
    category_filter = request.GET.get('category', '')
    min_price_filter = request.GET.get('min_price_filter', '')
    max_price_filter = request.GET.get('max_price_filter', '')

    # Definimos los parámetros para filtrar productos
    params = {
        'name': name_filter,
        'category': category_filter,
        'min_price_filter': min_price_filter,
        'max_price_filter': max_price_filter,
    }

    # Obtenemos los productos desde la API aplicando los filtros
    response = requests.get(settings.API_BASE_URL + 'product/', params=params)
    products = response.json()
    # Filtrar los productos para excluir los que tienen is_rentable=True
    filtered_products = [
        product for product in products if not product['is_rentable']]

    # Obtenemos las categorías desde la API
    categories = requests.get(settings.API_BASE_URL + 'category/').json()

    # Para limpiar los filtros
    if 'clear_filters' in request.GET:
        response = requests.get(settings.API_BASE_URL + 'product/').json()
        products = response

    data = {
        'products': filtered_products,
        'categories': categories,
    }

    return render(request, 'app/catalogue.html', data)

@api_view(['GET','POST']) 
def rental_service(request):
    if request.method == 'POST':
        form = RentalOrderForm(request.POST)
        if form.is_valid():
            # Obtener el objeto datetime del formulario
            deliver_date = form.cleaned_data['deliver_date']
            # Convertir el objeto datetime a una cadena de texto en formato ISO 8601
            deliver_date_iso = deliver_date.strftime('%Y-%m-%dT%H:%M')

            rental_order_data = {
                'rut': form.cleaned_data['rut'],
                'name': form.cleaned_data['name'],
                'address': form.cleaned_data['address'],
                'email': form.cleaned_data['email'],
                'phone': form.cleaned_data['phone'],
                # Utilizar la cadena de texto en lugar del objeto datetime
                'deliver_date': deliver_date_iso,
            }
            # validacion para verificar que el mismo rut no haya generado una solicitud de arriendo en los ultimos 15 minutos
            existing_order = RentalOrder.objects.filter(
                Q(rut=rental_order_data['rut']) &
                Q(created_at__gte=timezone.now() - timedelta(minutes=15))
            ).exists()

            if existing_order:
                return JsonResponse({'error': 'Ya existe una orden de renta con los mismos datos'})
            else:
                try:
                    # Crear la orden a través de la API
                    rental_order_response = requests.post(
                        settings.API_BASE_URL + 'rental-orders/', json=rental_order_data)
                    if rental_order_response.status_code == 201:
                        rental_order = rental_order_response.json()
                        # Obtener el ID de la orden de renta creada
                        rental_order_id = rental_order['id']

                        # Obtener la lista de productos seleccionados y sus cantidades
                        products_selected = request.POST.getlist('products')
                        quantities = [int(request.POST.get(
                            f'quantity_{product_id}', 1)) for product_id in products_selected]

                        # Obtener la lista completa de productos utilizando los ID de los productos seleccionados
                        products = Product.objects.filter(
                            id__in=products_selected)

                        rental_order_items = []
                        for product_id, quantity in zip(products_selected, quantities):
                            # Obtener el producto de la base de datos
                            product = Product.objects.get(id=product_id)
                            rental_order_item = RentalOrderItem(
                                rental_order=RentalOrder.objects.get(
                                    id=rental_order_id),
                                product_name=product.name,
                                product_price=product.price,
                                amount=quantity
                            )
                            rental_order_items.append(rental_order_item)

                        # Especificar un tamaño de lote adecuado
                        RentalOrderItem.objects.bulk_create(
                            rental_order_items, batch_size=100)

                        # Obtener los nombres de los productos y calcular el precio total
                        product_names = [product.name for product in products]
                        total_price = sum(
                            product.price * quantity for product, quantity in zip(products, quantities))

                        # Enviar correo electrónico con la información del RentalOrder
                        email_subject = 'Confirmación de orden de renta'
                        email_body = f'Se ha creado una nueva orden de renta con los siguientes detalles:\n\n' \
                                     f'RUT: {rental_order_data["rut"]}\n' \
                                     f'Nombre: {rental_order_data["name"]}\n' \
                                     f'Dirección: {rental_order_data["address"]}\n' \
                                     f'Correo electrónico: {rental_order_data["email"]}\n' \
                                     f'Teléfono: {rental_order_data["phone"]}\n' \
                                     f'Fecha de entrega: {rental_order_data["deliver_date"]}\n\n' \
                                     f'Productos:\n'
                        for name, quantity, product in zip(product_names, quantities, products):
                            email_body += f'- {name}: {quantity}\nprecio: {product.price}\n'
                        email_body += f'Total de la orden: {total_price}\n\n' \
                                      f'Gracias por su solicitud.'

                        sender_email = settings.EMAIL_HOST_USER
                        receiver_email = rental_order_data['email']

                        send_mail(email_subject, email_body,
                                  sender_email, [receiver_email])

                        return JsonResponse({'message': 'La solicitud de arriendo ha sido enviada correctamente, recibiras un correo con la información'})
                    else:
                        return JsonResponse({'error': 'Error al enviar la solicitud'})

                except Exception as e:
                    print(f"Error en el servidor: {str(e)}")
                    return JsonResponse({'error': 'Error en el servidor'})

        else:
            return JsonResponse({'error': 'Error en los datos del formulario'})

    else:
        form = RentalOrderForm()

    # Definimos los parámetros para filtrar productos
    params = {
        'is_rentable': 'true',
    }

    # Obtenemos los productos desde la API aplicando los filtros
    product_response = requests.get(
        settings.API_BASE_URL + 'product/', params=params).json()

    data = {
        'form': form,
        'products': product_response,
        'csrf_token': get_token(request)
    }

    return render(request, 'app/rental_service.html', data)

# CONTATO
@api_view(['GET','POST'])
def contact(request):
    data = {
        'form': ContactForm()
    }

    return render(request, 'app/contact/contact.html', data)

@require_http_methods(["POST"])
def update_contact_status(request, contact_id):
    if request.method == 'POST':
        status = request.POST.get('status')
        data = {
            'status': status
        }
        response = requests.patch(
            settings.API_BASE_URL + f'contact/{contact_id}/', data=data)

    return redirect('list_contact')

@user_passes_test(is_admin)
@require_http_methods(["GET"])
def list_contact(request):
    status = request.GET.get('status', '')
    query_type = request.GET.get('query_type', '')

    response_query_types = requests.get(settings.API_BASE_URL + 'query-type/')
    query_types = response_query_types.json()

    params = {}
    if status and status != 'Todos':
        params['status'] = status
    if query_type and query_type != 'Todos':
        params['query_type'] = query_type

    response = requests.get(settings.API_BASE_URL + 'contact/', params=params)
    if response.status_code == 200:
        contacts = response.json()
    else:
        contacts = []

    # Filtrar los contactos localmente en función del tipo de contacto seleccionado
    if query_type and query_type != 'Todos':
        contacts = [
            contact for contact in contacts if contact['query_type_name'] == query_type]

    paginator = Paginator(contacts, 5)
    page = request.GET.get('page')

    try:
        contacts = paginator.page(page)
    except PageNotAnInteger:
        contacts = paginator.page(1)
    except EmptyPage:
        contacts = paginator.page(paginator.num_pages)

    data = {
        'entity': contacts,
        'paginator': paginator,
        'query_types': query_types,
        'selected_status': status,
        'selected_query_type': query_type,
    }

    return render(request, 'app/contact/list.html', data)

# VISTAS DE QUERYTYPE

def get_object_query_type(id):
    response = requests.get(settings.API_BASE_URL + f'query-type/{id}/')

    if response.status_code == 200:
        query_type_data = response.json()
        return query_type_data
    else:
        print(f'Error al obtener el tipo de consulta: {response.content}')
        return None

@require_http_methods(["GET", "POST"])
def add_query_type(request):
    if request.method == 'POST':
        form = QueryTypeForm(request.POST, request.FILES)
        if form.is_valid():
            error_message = None  # Inicializamos la variable error_message
            try:
                serializer = QueryTypeSerializer(data=form.cleaned_data)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                messages.success(
                    request, 'Tipo de consulta agregada exitosamente.')
                return redirect('list_query_type')
            except serializers.ValidationError as e:
                # Agregar mensaje de error al campo 'name'
                form.add_error('name', e.detail['name'][0])
        else:
            error_message = "Error en los datos del formulario"
        data = {
            'form': form,
            'error_message': error_message
        }
    else:
        data = {
            'form': QueryTypeForm()
        }

    return render(request, 'app/querytype/add.html', data)

@user_passes_test(is_admin)
@require_http_methods(["GET"])
def list_query_type(request):
    response = requests.get(settings.API_BASE_URL + 'query-type/')
    query_types = response.json()
    page = request.GET.get('page', 1)

    try:
        paginator = Paginator(query_types, 5)
        query_types = paginator.page(page)
    except:
        raise Http404

    data = {
        'entity': query_types,
        'paginator': paginator
    }
    return render(request, 'app/querytype/list.html', data)

@require_http_methods(["GET", "POST"])
def update_query_type(request, id):
    querytype_data = get_object_query_type(id)

    if querytype_data:
        error_message = ""

        if request.method == 'POST':
            form = QueryTypeForm(request.POST)

            if form.is_valid():
                name = form.cleaned_data['name']

                # Verificar si existe una categoría con el mismo nombre a través de la API
                response = requests.get(
                    settings.API_BASE_URL + f'query-type/?name={name}')

                if response.status_code == 200:
                    existing_querytype = response.json()

                    if existing_querytype:
                        # Verificar si alguna categoría tiene un nombre diferente al nombre actual
                        for existing_querytype in existing_querytype:
                            if existing_querytype['name'] == name and existing_querytype['id'] != int(id):
                                form.add_error(
                                    'name', 'Este tipo de consulta ya existe')
                                error_message = "Este tipo de consulta ya existe"
                                print(
                                    "existing_querytype['name']: ", existing_querytype['name'])
                                print("name: ", name)
                                break  # Salir del bucle si se encuentra una categoría existente

                    if not error_message:
                        description = form.cleaned_data['description']

                        # Crear un nuevo diccionario con los datos actualizados
                        updated_data = {
                            'name': name,
                            'description': description
                        }

                        # Actualizar la categoría a través de la API
                        update_url = settings.API_BASE_URL + \
                            f'query-type/{id}/'
                        response = requests.put(update_url, data=updated_data)

                        if response.status_code == 200:
                            print('Tipo de consulta actualizado exitosamente')
                            messages.success(
                                request, "Modificado correctamente")
                            return redirect(to="list_query_type")
                        else:
                            print(
                                f'Error al actualizar el tipo de consulta: {response.content}')
                            error_message = "Error al actualizar el tipo de consulta a través de la API"
                else:
                    print(
                        f'Error al verificar la existencia de el tipo de consulta: {response.content}')
                    error_message = "Error al verificar la existencia de el tipo de consulta a través de la API"
            else:
                error_message = "Error en los datos del formulario"
        else:
            form = QueryTypeForm(initial=querytype_data)
            error_message = ""

        data = {
            'form': form,
            'error_message': error_message
        }

        return render(request, 'app/querytype/update.html', data)
    else:
        # Manejar el caso si no se puede obtener la categoría de la API
        messages.error(
            request, "Error al obtener el tipo de consulta de la API")
        return redirect(to="list_query_type")

@require_http_methods(["POST"])
def delete_query_type(request, id):
    querytype_data = get_object_query_type(id)

    if querytype_data:
        # Crear una instancia de Product solo con el ID
        querytype = QueryType(id=querytype_data['id'])

        # Realizar una solicitud DELETE a la API para eliminar el producto
        delete_response = requests.delete(
            settings.API_BASE_URL + f'query-type/{id}/')

        if delete_response.status_code == 204:
            querytype.delete()
            messages.success(request, "Eliminado correctamente")
            return redirect(to="list_query_type")
        else:
            # Manejar el caso de error en la solicitud DELETE
            print(
                f'Error al eliminar el tipo de consulta: {delete_response.content}')
            error_message = "Error al eliminar el tipo de consulta a través de la API"
            data = {
                'form': QueryTypeForm(instance=querytype),
                'error_message': error_message
            }
            return render(request, 'app/query-type/update.html', data)
    else:
        # Manejar el caso de error al obtener el producto
        error_message = "Error al obtener el tipo de consulta a través de la API"
        data = {
            'error_message': error_message
        }
        return render(request, 'app/query-type/update.html', data)

# VISTAS DE PRODUCT


def get_object_product(id):
    response = requests.get(settings.API_BASE_URL + f'product/{id}/')

    if response.status_code == 200:
        product_data = response.json()
        product_data.pop('image', None)  # Eliminar el campo de imagen del JSON
        return product_data
    else:
        print(f'Error al obtener el producto: {response.content}')
        return None


@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def add_product(request):
    if request.method == 'POST':
        error_message = ""
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            name = form.cleaned_data['name']

            # Verificar si existe un producto con el mismo nombre a través de la API
            response = requests.get(
                settings.API_BASE_URL + f'product/?name={name}')

            if response.status_code == 200:
                existing_products = response.json()

                if existing_products:
                    # Verificar si algún producto tiene un nombre diferente al nombre actual
                    for existing_product in existing_products:
                        if existing_product['name'] == name and existing_product['id'] != id:
                            form.add_error('name', 'Este producto ya existe')
                            error_message = "Este producto ya existe"
                            print(
                                "existing_product['name']: ", existing_product['name'])
                            print("name: ", name)
                            break  # Salir del bucle si se encuentra un producto existente
                if not error_message:
                    price = form.cleaned_data['price']
                    description = form.cleaned_data['description']
                    is_new = form.cleaned_data['is_new']
                    category_id = form.cleaned_data['category'].id
                    stock = form.cleaned_data['stock']
                    is_featured = form.cleaned_data['is_featured']
                    image = form.cleaned_data['image']
                    is_rentable = form.cleaned_data['is_rentable']

                    product_data = {
                        'name': name,
                        'price': price,
                        'description': description,
                        'is_new': is_new,
                        'category': category_id,
                        'stock': stock,
                        'is_featured': is_featured,
                        'is_rentable': is_rentable,
                    }

                    response = requests.post(
                        settings.API_BASE_URL + 'product/',
                        data=product_data,
                        files={'image': image}
                    )

                    if response.status_code == 201:
                        print('Producto creado exitosamente')
                        messages.success(
                            request, 'Producto agregado exitosamente.')
                        return redirect('list_product')
                    else:
                        print(
                            f'Error al crear el producto: {response.content}')
                        error_message = "Error al crear el producto a través de la API"
            else:
                print(
                    f'Error al verificar la existencia del producto: {response.content}')
                error_message = "Error al verificar la existencia del producto a través de la API"
        else:
            error_message = "Error en los datos del formulario"
        data = {
            'form': form,
            'error_message': error_message
        }
    else:
        data = {
            'form': ProductForm()
        }
    return render(request, 'app/product/add.html', data)


@user_passes_test(is_admin)
@require_http_methods(["GET"])
def list_product(request):
    name_filter = request.GET.get('name', '')
    category_filter = request.GET.get('category', '')

    params = {}
    if name_filter:
        params['name'] = name_filter
    if category_filter:
        params['category'] = category_filter

    response = requests.get(settings.API_BASE_URL + 'product/', params=params)
    if response.status_code == 200:
        products = response.json()
    else:
        products = []

    paginator = Paginator(products, 5)
    page = request.GET.get('page')

    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    response = requests.get(settings.API_BASE_URL + 'category/')
    categories = response.json()

    data = {
        'entity': products,
        'paginator': paginator,
        'name_filter': name_filter,
        'category_filter': category_filter,
        'categories': categories,
    }
    return render(request, 'app/product/list.html', data)


@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def update_product(request, id):
    product_data = get_object_product(id)

    if product_data:
        error_message = ""

        if request.method == 'POST':
            form = ProductForm(request.POST, request.FILES)

            if form.is_valid():
                name = form.cleaned_data['name']

                # Verificar si existe un producto con el mismo nombre a través de la API
                response = requests.get(
                    settings.API_BASE_URL + f'product/?name={name}')

                if response.status_code == 200:
                    existing_products = response.json()

                    if existing_products:
                        # Verificar si algún producto tiene un nombre diferente al nombre actual
                        for existing_product in existing_products:
                            if existing_product['name'] == name and existing_product['id'] != id:
                                form.add_error(
                                    'name', 'Este producto ya existe')
                                error_message = "Este producto ya existe"
                                print(
                                    "existing_product['name']: ", existing_product['name'])
                                print("name: ", name)
                                break  # Salir del bucle si se encuentra un producto existente

                    if not error_message:
                        description = form.cleaned_data['description']
                        price = form.cleaned_data['price']
                        is_new = form.cleaned_data['is_new']
                        category_id = form.cleaned_data['category'].id
                        stock = form.cleaned_data['stock']
                        is_featured = form.cleaned_data['is_featured']
                        image = form.cleaned_data['image']
                        is_rentable = form.cleaned_data['is_rentable']

                        # Crear un nuevo diccionario con los datos actualizados
                        updated_data = {
                            'name': name,
                            'description': description,
                            'price': price,
                            'is_new': is_new,
                            'category': category_id,
                            'stock': stock,
                            'is_featured': is_featured,
                            'is_rentable': is_rentable
                        }

                        # Actualizar el producto a través de la API
                        update_url = settings.API_BASE_URL + f'product/{id}/'
                        files = {'image': image}  # Archivo adjunto
                        response = requests.put(
                            update_url, data=updated_data, files=files)

                        if response.status_code == 200:
                            print('Producto actualizado exitosamente')
                            messages.success(
                                request, "Modificado correctamente")
                            return redirect(to="list_product")
                        else:
                            print(
                                f'Error al actualizar el producto: {response.content}')
                            error_message = "Error al actualizar el producto a través de la API"
                else:
                    print(
                        f'Error al verificar la existencia del producto: {response.content}')
                    error_message = "Error al verificar la existencia del producto a través de la API"
            else:
                error_message = "Error en los datos del formulario"
        else:
            form = ProductForm(initial=product_data)
            error_message = ""

        data = {
            'form': form,
            'error_message': error_message
        }

        return render(request, 'app/product/update.html', data)
    else:
        # Manejar el caso si no se puede obtener el producto de la API
        messages.error(request, "Error al obtener el producto de la API")
        return redirect(to="list_product")


@user_passes_test(is_admin)
@require_http_methods(["GET"])
def delete_product(request, id):
    product_data = get_object_product(id)

    if product_data:
        # Crear una instancia de Product solo con el ID
        product = Product(id=product_data['id'])

        # Realizar una solicitud DELETE a la API para eliminar el producto
        delete_response = requests.delete(
            settings.API_BASE_URL + f'product/{id}/')

        if delete_response.status_code == 204:
            product.delete()
            messages.success(request, "Eliminado correctamente")
            return redirect(to="list_product")
        else:
            # Manejar el caso de error en la solicitud DELETE
            print(f'Error al eliminar el producto: {delete_response.content}')
            error_message = "Error al eliminar el producto a través de la API"
            messages.error(request, error_message)
            # Redireccionar a la página de listado con mensaje de error
            return redirect(to="list_product")
    else:
        # Manejar el caso de error al obtener el producto
        error_message = "Error al obtener el producto a través de la API"
        messages.error(request, error_message)
        # Redireccionar a la página de listado con mensaje de error
        return redirect(to="list_product")

@require_http_methods(["GET"])
def product_detail(request, id):
    # Realizar una solicitud GET a la API para obtener los detalles del producto
    response = requests.get(settings.API_BASE_URL + f'product/{id}/')

    if response.status_code == 200:
        product_data = response.json()

        # Obtener la instancia de Category a través de la API
        category_id = product_data['category']
        category_data = get_object_category(category_id)

        if category_data:
            # Crear el objeto Category con los datos obtenidos de la API
            category = Category(**category_data)

            # Remover el campo 'category_name' del diccionario product_data
            product_data.pop('category_name', None)

            # Actualizar el campo 'category' en product_data con la instancia de Category
            product_data['category'] = category

            # Crear el objeto Product con los datos actualizados
            product = Product(**product_data)

            data = {
                'product': product
            }
            return render(request, 'app/product/detail.html', data)
        else:
            # Manejar el caso si no se puede obtener la categoría de la API
            error_message = "Error al obtener la categoría de la API"
            return render(request, 'app/product/detail.html', {'error_message': error_message})
    else:
        # Manejar el caso de error en la solicitud
        print(
            f'Error al obtener los detalles del producto: {response.content}')
        error_message = "Error al obtener los detalles del producto a través de la API"
        return render(request, 'app/product/detail.html', {'error_message': error_message})

# VISTA DE REGISTRO NO API


# METODOS DEL CARRITO NO API
@require_http_methods(["GET"])
def add_prod_cart(request, product_id):
    cart = Cart(request)
    product = Product.objects.get(id=product_id)

    if product.stock <= 0:
        messages.error(request, "Error: Product is out of stock.")
    elif cart.get_product_quantity(product) >= product.stock:
        messages.error(request, "Error: Maximum stock limit reached.")
    else:
        cart.add(product)
        # messages.success(request, "Product added to cart successfully.")

    return redirect(to="Cart")


@require_http_methods(["GET"])
def del_prod_cart(request, product_id):
    cart = Cart(request)
    product = Product.objects.get(id=product_id)
    cart.delete(product)
    return redirect(to="Cart")


@require_http_methods(["GET"])
def subtract_product_cart(request, product_id):
    cart = Cart(request)
    product = Product.objects.get(id=product_id)
    cart.subtract(product)
    return redirect("Cart")

@require_http_methods(["GET"])
def clean_cart(request):
    cart = Cart(request)
    cart.clean()
    return redirect("Cart")

@require_http_methods(["GET"])
def cart_page(request):
    products = Product.objects.all()
    data = {
        'products': products
    }

    return render(request, 'app/cart_page.html', data)

@require_http_methods(["GET", "POST"])
def buy_confirm(request):
    cart = Cart(request)
    cart.buy()
    cart.clean()
    messages.success(request, "Compra de prueba Completada")
    return redirect('home')

# VISTAS CATEGORY


def get_object_category(id):
    response = requests.get(settings.API_BASE_URL + f'category/{id}/')

    if response.status_code == 200:
        category_data = response.json()
        category_data.pop('image', None)
        return category_data
    else:
        print(f'Error al obtener la categoria: {response.content}')
        return None


@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            error_message = None  # Inicializamos la variable error_message
            try:
                serializer = CategorySerializer(data=form.cleaned_data)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                messages.success(request, 'Categoría agregada exitosamente.')
                return redirect('list_category')
            except serializers.ValidationError as e:
                # Agregar mensaje de error al campo 'name'
                form.add_error('name', e.detail['name'][0])
        else:
            error_message = "Error en los datos del formulario"
        data = {
            'form': form,
            'error_message': error_message
        }
    else:
        data = {
            'form': CategoryForm()
        }

    return render(request, 'app/category/add.html', data)


@user_passes_test(is_admin)
@require_http_methods(["GET"])
def list_category(request):
    response = requests.get(settings.API_BASE_URL + 'category/')
    categories = response.json()
    page = request.GET.get('page', 1)

    try:
        paginator = Paginator(categories, 5)
        categories = paginator.page(page)
    except:
        raise Http404

    data = {
        'entity': categories,
        'paginator': paginator
    }
    return render(request, 'app/category/list.html', data)


@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def update_category(request, id):
    category_data = get_object_category(id)

    if category_data:
        error_message = ""

        if request.method == 'POST':
            form = CategoryForm(request.POST, request.FILES)

            if form.is_valid():
                name = form.cleaned_data['name']

                # Verificar si existe una categoría con el mismo nombre a través de la API
                response = requests.get(
                    settings.API_BASE_URL + f'category/?name={name}')

                if response.status_code == 200:
                    existing_categories = response.json()

                    if existing_categories:
                        # Verificar si alguna categoría tiene un nombre diferente al nombre actual
                        for existing_category in existing_categories:
                            if existing_category['name'] == name and existing_category['id'] != int(id):
                                form.add_error(
                                    'name', 'Esta categoría ya existe')
                                error_message = "Esta categoría ya existe"
                                print(
                                    "existing_category['name']: ", existing_category['name'])
                                print("name: ", name)
                                break  # Salir del bucle si se encuentra una categoría existente

                    if not error_message:
                        description = form.cleaned_data['description']
                        image = form.cleaned_data['image']

                        # Crear un nuevo diccionario con los datos actualizados
                        updated_data = {
                            'name': name,
                            'description': description
                        }

                        # Actualizar la categoría a través de la API
                        update_url = settings.API_BASE_URL + f'category/{id}/'
                        files = {'image': image}  # Archivo adjunto
                        response = requests.put(
                            update_url, data=updated_data, files=files)

                        if response.status_code == 200:
                            print('Categoría actualizada exitosamente')
                            messages.success(
                                request, "Modificado correctamente")
                            return redirect(to="list_category")
                        else:
                            print(
                                f'Error al actualizar la categoría: {response.content}')
                            error_message = "Error al actualizar la categoría a través de la API"
                else:
                    print(
                        f'Error al verificar la existencia de la categoría: {response.content}')
                    error_message = "Error al verificar la existencia de la categoría a través de la API"
            else:
                error_message = "Error en los datos del formulario"
        else:
            form = CategoryForm(initial=category_data)
            error_message = ""

        data = {
            'form': form,
            'error_message': error_message
        }

        return render(request, 'app/category/update.html', data)
    else:
        # Manejar el caso si no se puede obtener la categoría de la API
        messages.error(request, "Error al obtener la categoría de la API")
        return redirect(to="list_category")


@user_passes_test(is_admin)
@require_http_methods(["GET"])
def delete_category(request, id):
    category_data = get_object_category(id)

    if category_data:
        # Crear una instancia de Product solo con el ID
        category = Category(id=category_data['id'])

        # Realizar una solicitud DELETE a la API para eliminar el producto
        delete_response = requests.delete(
            settings.API_BASE_URL + f'category/{id}/')

        if delete_response.status_code == 204:
            category.delete()
            messages.success(request, "Eliminado correctamente")
            return redirect(to="list_category")
        else:
            # Manejar el caso de error en la solicitud DELETE
            print(f'Error al eliminar la categoria: {delete_response.content}')
            error_message = "Error al eliminar la categoria a través de la API"
            data = {
                'form': CategoryForm(instance=category),
                'error_message': error_message
            }
            return render(request, 'app/category/update.html', data)
    else:
        # Manejar el caso de error al obtener el producto
        error_message = "Error al obtener la categoria a través de la API"
        data = {
            'error_message': error_message
        }
        return render(request, 'app/category/update.html', data)

# PANEL DE ADMIN

@user_passes_test(is_admin)
@require_http_methods(["GET", "POST"])
def admin_panel(request):

    return render(request, 'app/admin_panel.html')

@require_http_methods(["GET", "POST"])
def checkout_view(request):
    # Obtén las regiones y comunas desde la API
    regiones = LocationAPI.get_regions()
    comunas = LocationAPI.get_municipalities()
    
    # Calcula el cart_total sumando los precios y cantidades de los productos en el carrito
    cart_items = request.session.get('cart', {}).items()
    cart_total = sum(value.get('product_price', 0) * value.get('amount', 1) for key, value in cart_items)

    # Prepara las opciones de región para pasarlas al formulario
    region_choices = [('', 'Seleccione una región')] + [(region['id'], region['name']) for region in regiones]
    municipality_choices = [('', 'Seleccione una comuna')] + [(comuna['id'], comuna['name']) for comuna in comunas]

    # Inicializa el formulario con datos iniciales y opciones dinámicas
    initial_data = {
        "user": request.user.id if request.user.is_authenticated else None,
        "accumulated": cart_total,
    }
    form = OrderForm(initial=initial_data, region_choices=region_choices, municipality_choices=municipality_choices)

    context = {
        "form": form,
        "regiones": regiones,
        "comunas": comunas,
        "cart_total": cart_total,
    }
    return render(request, "app/checkout.html", context)

@require_http_methods(["GET", "POST"])
def user_login(request):
    global tok
    datos = {
        'form': LoginForm()
    }
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            usernameU = request.POST['usrN']
            passwordU = request.POST['pswrdN']
            user = authenticate(username=usernameU, password=passwordU)
            if user is not None:
                login(request, user)
                refresh = RefreshToken.for_user(user)
                token = str(refresh.access_token)

                # Guardar el token en el modelo Tokens
                token_data = {
                    'token': token,
                    'user': usernameU
                }
                token_serializer = TokenSerializer(data=token_data)
                if token_serializer.is_valid():
                    token_serializer.save()

                tok = token
                messages.success(request, "has iniciado sesión")
                return redirect(to="home")
    return render(request, "registration/login.html", datos)

@require_http_methods(["GET", "POST"])
def Recuperar(request):
    form = RecuperarForm(request.POST or None)
    if form.is_valid():
        email = form.cleaned_data['email']
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            user = None
        if user:
            # Generar una nueva contraseña temporal
            new_password = get_random_string(length=12)

            # Actualizar la contraseña del usuario
            user.set_password(new_password)
            user.save()

            # Enviar la nueva contraseña por correo
            subject = 'Recuperación de contraseña'
            message = f'Tu nueva contraseña temporal es: {new_password}. Por favor, cámbiala inmediatamente después de iniciar sesión.'
            from_email = settings.EMAIL_HOST_USER
            recipient_list = [email]
            send_mail(subject, message, from_email,
                      recipient_list, fail_silently=False)

            # Mostrar mensaje de éxito
            messages.success(
                request, "Se ha enviado una nueva contraseña temporal a tu correo.")
            # Redirige al formulario de login después de enviar el correo
            return redirect('login')

        else:
            messages.error(
                request, 'No se encontró una cuenta con ese correo electrónico.')

    return render(request, 'registration/Recuperar.html', {'form': form})

# se crea usuario nuevo y token

class LoginView(APIView):
    def post(self, request, format=None):
        django_request = HttpRequest()
        django_request.method = request.method
        django_request.POST = request.data
        django_request._request = request._request

        serializer = LoginSerializer(data=django_request.POST)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        return Response({'token': token.key})

@require_http_methods(["GET", "POST"])
def Registrar(request):
    datos = {
        'form': UsuariosForm()
    }
    if request.method == 'POST':
        form = UsuariosForm(request.POST)
        if form.is_valid():
            # Obtiene los datos del usuario desde el formulario
            usernameN = form.cleaned_data.get('usrN')
            passwordN = form.cleaned_data.get('pswrdN')
            passwordN2 = form.cleaned_data.get('pswrdN2')
            try:
                # Verifica la existencia del usuario
                user = User.objects.get(username=usernameN)
            except User.DoesNotExist:
                # Si no existe, se genera un nuevo usuario validando si las contraseñas son idénticas
                if passwordN == passwordN2:
                    user = User.objects.create_user(
                        username=usernameN, email=usernameN, password=passwordN)
                    # Autentifica las credenciales del usuario
                    user = authenticate(username=usernameN, password=passwordN)
                    if user is not None:
                        login(request, user)
                        refresh = RefreshToken.for_user(user)
                        token = str(refresh.access_token)

                        # Guardar el token en el modelo Tokens
                        token_data = {
                            'token': token,
                            'user': usernameN
                        }
                        token_instance = Tokens.objects.create(**token_data)

                        messages.success(
                            request, "Te has registrado correctamente")
                        return redirect('home')

    return render(request, "registration/Registrar.html", datos)

@require_http_methods(["GET", "POST"])
def update_last_order_paid_status(user):
    try:
        last_order = Order.objects.filter(user=user).latest('id')
        last_order.pagado = True
        last_order.save()
    except Order.DoesNotExist:
        pass

@require_http_methods(["GET", "POST"])
def payment_success(request):
    if request.method == 'POST':
        user = request.user if request.user.is_authenticated else None
        name = request.POST.get('name')
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        accumulated = request.POST.get('accumulated')

        # Crear la orden
        order = Order(user=user, name=name, address=address, phone=phone, accumulated=accumulated)
        
        # Guarda la orden y asegúrate de que el ID se genere
        order.save()

        # Agregar los productos a la orden
        cart_items = request.session.get('cart', {}).items()
        for key, value in cart_items:
            product_name = value.get('product_name')
            product_price = value.get('product_price')
            amount = value.get('amount')
            order_item = OrderItem(order=order, product_name=product_name, product_price=product_price, amount=amount)
            order_item.save()

        # Asegúrate de pasar el objeto `order` al template
        return render(request, 'app/payment_success.html', {'order': order})

    return render(request, 'app/checkout.html')

@require_http_methods(["GET", "POST"])

def order_list(request):
    # Obtener todas las órdenes con prefetch para optimización
    orders = Order.objects.prefetch_related('orderitem_set').all()
    page = request.GET.get('page', 1)

    # Filtros por rango de fecha
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        orders = orders.filter(fecha__range=[start_date, end_date])

    # Filtro por nombre de OrderItem
    order_item_name = request.GET.get('order_item_name')
    if order_item_name:
        orders = orders.filter(orderitem__product_name__icontains=order_item_name)

    # Cálculos generales
    total_orders_count = orders.count()  # Contar el total de órdenes
        
    total_accumulated = orders.aggregate(total_accumulated=Sum('accumulated'))['total_accumulated'] or 0
    total_products_sold = OrderItem.objects.filter(order__in=orders).aggregate(
        total_sold=Sum('amount')
    )['total_sold'] or 0

    # Top productos más vendidos
    top_products = OrderItem.objects.filter(order__in=orders).values(
        'product_name'
    ).annotate(
        total_amount=Sum('amount')
    ).order_by('-total_amount')[:4]

    # Dividir órdenes pagadas y no pagadas
    paid_orders_count = orders.filter(pagado=True).count()
    unpaid_orders_count = orders.filter(pagado=False).count()

    # Ventas mensuales
    monthly_sales = [
        orders.filter(fecha__month=month).count() for month in range(1, 13)
    ]

    # Totales por usuario (top 10 usuarios con mayores acumulados)
    user_totals = Order.objects.values('user__username').annotate(
        total_accumulated=Sum('accumulated')
    ).order_by('-total_accumulated')[:10]

    # Paginación
    paginator = Paginator(orders, 5)
    try:
        orders = paginator.page(page)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)

    # Contexto para la plantilla
    context = {
        'entity': orders,
        'paginator': paginator,
        'total_accumulated': total_accumulated,
        'total_orders_count': total_orders_count,  # Incluyendo el total de órdenes
        'total_products_sold': total_products_sold,
        'top_products': top_products,
        'monthly_sales': monthly_sales,
        'month_labels': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'],
        'paid_orders_count': paid_orders_count,
        'unpaid_orders_count': unpaid_orders_count,
        'user_totals': user_totals,  # Añadir los totales por usuario al contexto
    }
    return render(request, 'app/order_list.html', context)


@api_view(['POST'])
def obtain_token(request):
    username = request.data.get('username')
    password = request.data.get('password')

    if username and password:
        user = authenticate(username=username, password=password)
        if user is not None:
            refresh = RefreshToken.for_user(user)
            return Response({
                'access_token': str(refresh.access_token),
                'refresh_token': str(refresh),
            })
    return Response({'error': 'Credenciales inválidas.'}, status=400)

@require_http_methods(["GET", "POST"])
def list_rental_order(request):
    product_name = request.GET.get('product_name')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    response = requests.get(settings.API_BASE_URL + 'rental-orders/')
    if response.status_code != 200:
        error_message = 'Error al obtener los datos de la API'
        return HttpResponse(error_message, status=500)

    rental_orders = response.json()

    # Aplicar los filtros después de recibir los datos de la API
    if product_name:
        rental_orders = [ro for ro in rental_orders if any(
            product_name.lower() in item['product_name'].lower() for item in ro['items'])]

    if start_date and end_date:
        start_date = parse(start_date).date()
        end_date = parse(end_date).date()
        rental_orders = [ro for ro in rental_orders if start_date <= parse(
            ro['deliver_date']).date() <= end_date]

    # Calcular el precio acumulado y la cantidad de productos vendidos
    total_accumulated = sum(
        float(item['product_price']) * item['amount']
        for ro in rental_orders
        for item in ro['items']
    )
    total_products_sold = sum(
        item['amount']
        for ro in rental_orders
        for item in ro['items']
    )

    # Calcular el campo total_price para cada rental_order
    for rental_order in rental_orders:
        total_price = 0
        for item in rental_order['items']:
            product_price = float(item['product_price'])
            amount = item['amount']
            total_price += product_price * amount
        rental_order['total_price'] = total_price

    # Obtener una lista de todos los productos vendidos
    all_products = [
        item['product_name']
        for ro in rental_orders
        for item in ro['items']
    ]

    # Contar la cantidad de veces que se vende cada producto
    product_counts = Counter(all_products)

    # Obtener los productos más vendidos
    top_products = []
    for product, _ in product_counts.most_common(4):
        total_amount = sum(
            item['amount']
            for ro in rental_orders
            for item in ro['items']
            if item['product_name'] == product
        )
        top_products.append((product, total_amount))

    # Crear un Paginator con los datos sin paginar
    paginator = Paginator(rental_orders, 5)
    page = request.GET.get('page', 1)

    try:
        rental_orders = paginator.page(page)
    except:
        error_message = 'Error al paginar los datos'
        return HttpResponse(error_message, status=500)

    data = {
        'entity': rental_orders,
        'paginator': paginator,
        'total_accumulated': total_accumulated,
        'total_products_sold': total_products_sold,
        'top_products': top_products,
    }
    return render(request, "app/rental_order/list.html", data)


# Implementacion API Transbank
def webpay_init_transaction(request):
    if request.method == 'POST':
        buy_order = request.POST.get("buy_order")
        session_id = request.POST.get("session_id")
        amount = request.POST.get("amount")

        # Verificar si los valores están vacíos
        if not buy_order or not amount or not session_id:
            return JsonResponse({"error": "Datos faltantes en la solicitud de transacción"})

        # Configuración de Transbank para el ambiente de integración
        return_url = "http://127.0.0.1:8000/webpay/return/"
        tx = Transaction(
            WebpayOptions(
                "597055555532",                       
                "579B532A7440BB0C9079DED94D31EA1615BACEB56610332264630D42D0A36B1C",  
                IntegrationType.TEST              # Modo de prueba, cambiar a .LIVE para PRODUCCION    
            )
        )

        response = tx.create(buy_order, session_id, amount, return_url)

        if "token" in response and "url" in response:
            return redirect(response["url"] + "?token_ws=" + response["token"])
        else:
            return JsonResponse({"error": "Error al iniciar la transacción con Webpay.", "detalle": response})

    return JsonResponse({"error": "Método no permitido"}, status=405)

def webpay_return(request):
    """
    Confirma la transacción cuando el usuario regresa desde Webpay.
    """
    token = request.GET.get("token_ws")

    # Configuración de Transbank para el ambiente de integración
    tx = Transaction(
        WebpayOptions(
            "597055555532",                        # Código de comercio de prueba
            "579B532A7440BB0C9079DED94D31EA1615BACEB56610332264630D42D0A36B1C",  # API Key de prueba
            IntegrationType.TEST                   # Modo de prueba, cambiar a .LIVE para PRODUCCION
        )
    )

    # Realizar la confirmación de la transacción
    response = tx.commit(token)
    
    # Acceder a los valores del diccionario directamente
    if response.get("response_code") == 0:
        return render(request, "app/transbank/payment_success.html", {"response": response})
    else:
        return render(request, "app/transbank/payment_failed.html", {"error": "Transacción rechazada."})
    
# API VIEWS para Location
@api_view(['GET'])
def list_regions(request):
    regions = Region.objects.all()
    serializer = RegionSerializer(regions, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def list_municipalities(request):
    region_id = request.GET.get('region_id')
    if region_id:
        municipalities = Municipality.objects.filter(region_id=region_id)
    else:
        municipalities = Municipality.objects.all()
    
    serializer = MunicipalitySerializer(municipalities, many=True)
    return Response(serializer.data)



import openpyxl
from openpyxl.styles import Font, Alignment
from django.db.models import Count, Sum, F
from django.http import HttpResponse

def generate_excel_report(request):
    # Obtener las órdenes
    orders = Order.objects.prefetch_related('orderitem_set').all()

    # Calcular totales
    total_accumulated = orders.aggregate(total_accumulated=Sum('accumulated'))[
        'total_accumulated']
    total_products_sold = OrderItem.objects.aggregate(
        total_sold=Sum('amount'))['total_sold']

    # Resumen por usuario
    users_summary = orders.values('user__username').annotate(
        total_orders=Count('order_id'),
        total_amount=Sum('accumulated')
    )

    # Ventas mensuales
    months = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
              'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    monthly_sales = [orders.filter(fecha__month=i).aggregate(
        total=Sum('accumulated'))['total'] or 0 for i in range(1, 13)]

    # Estadísticas globales
    total_orders = orders.count()
    unique_users = orders.values('user').distinct().count()
    avg_orders_per_user = total_orders / unique_users if unique_users > 0 else 0

    # Crear un archivo Excel
    wb = openpyxl.Workbook()

    # **Hoja Principal: Detalles de las Órdenes**
    ws = wb.active
    ws.title = "Detalles de Órdenes"
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=12)
    align_center = Alignment(horizontal="center", vertical="center")

    # Encabezados
    ws.append(["Usuario", "Número de Orden", "Total",
              "Fecha", "Pagado", "Productos"])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = align_center

    # Agregar datos de las órdenes
    for order in orders:
        product_details = "\n".join(
            [f"{item.product_name} (x{item.amount})" for item in order.orderitem_set.all(
            )]
        )
        ws.append([
            order.user.username if order.user else "Anónimo",
            order.order_id,
            order.accumulated,
            order.fecha.strftime("%Y-%m-%d"),
            "Sí" if order.pagado else "No",
            product_details
        ])

    # Agregar una línea separadora
    ws.append([])

    # Título para estadísticas globales
    ws.append(["Estadísticas Generales"])
    ws[f"A{ws.max_row}"].font = title_font
    ws[f"A{ws.max_row}"].alignment = align_center

    ws.append(["Total de Órdenes", total_orders])
    ws.append(["Usuarios Únicos", unique_users])
    ws.append(["Promedio de Órdenes por Usuario",
              round(avg_orders_per_user, 2)])

    # ================================
    # Resumen General de Órdenes
    # ================================
    ws.append([])
    ws.append(["Resumen General de Órdenes"])
    ws[f"A{ws.max_row}"].font = title_font
    ws[f"A{ws.max_row}"].alignment = align_center

    ws.append(["Total Órdenes", "Órdenes Pagadas",
              "Órdenes No Pagadas", "Total Acumulado"])

    paid_orders = Order.objects.filter(pagado=True).count()
    unpaid_orders = Order.objects.filter(pagado=False).count()
    total_accumulated = Order.objects.aggregate(
        total=Sum('accumulated'))['total'] or 0
    ws.append([total_orders, paid_orders, unpaid_orders, total_accumulated])

    # ================================
    # Top Productos Más Vendidos
    # ================================
    ws.append([])
    ws.append(["Top Productos Más Vendidos"])
    ws[f"A{ws.max_row}"].font = title_font
    ws[f"A{ws.max_row}"].alignment = align_center

    ws.append(["Producto", "Cantidad Vendida"])
    top_products = OrderItem.objects.values('product_name').annotate(
        total_sold=Sum('amount')
    ).order_by('-total_sold')[:10]
    for product in top_products:
        ws.append([product['product_name'], product['total_sold']])

    # ================================
    # Top Clientes
    # ================================
    ws.append([])
    ws.append(["Top Clientes"])
    ws[f"A{ws.max_row}"].font = title_font
    ws[f"A{ws.max_row}"].alignment = align_center

    ws.append(["Usuario", "Total Órdenes", "Monto Acumulado"])
    top_clients = Order.objects.values('user__username').annotate(
        total_orders=Count('user_id'),
        total_accumulated=Sum('accumulated')
    ).order_by('-total_accumulated')[:10]
    for client in top_clients:
        ws.append([client['user__username'], client['total_orders'],
                  client['total_accumulated']])

    # ================================
    # Contactos por Estado
    # ================================
    ws.append([])
    ws.append(["Contactos por Estado"])
    ws[f"A{ws.max_row}"].font = title_font
    ws[f"A{ws.max_row}"].alignment = align_center

    ws.append(["Estado", "Cantidad"])
    contact_status_summary = Contact.objects.values('status').annotate(
        total=Count('id')
    )
    for status in contact_status_summary:
        ws.append([status['status'], status['total']])

    # ================================
    # Contactos por Tipo de Consulta
    # ================================
    ws.append([])
    ws.append(["Contactos por Tipo de Consulta"])
    ws[f"A{ws.max_row}"].font = title_font
    ws[f"A{ws.max_row}"].alignment = align_center

    ws.append(["Tipo de Consulta", "Cantidad"])
    contact_type_summary = Contact.objects.values('query_type__name').annotate(
        total=Count('id')
    )
    for query_type in contact_type_summary:
        ws.append([query_type['query_type__name'], query_type['total']])

    # ================================
    # Resumen de Rentas
    # ================================
    ws.append([])
    ws.append(["Resumen de Rentas"])
    ws[f"A{ws.max_row}"].font = title_font
    ws[f"A{ws.max_row}"].alignment = align_center

    ws.append(["Total Rentas", "Ingresos Totales"])
    total_rentals = RentalOrder.objects.count()
    total_rental_income = RentalOrderItem.objects.aggregate(
        total_income=Sum(F('product_price') * F('amount'))
    )['total_income'] or 0
    ws.append([total_rentals, total_rental_income])

    # ================================
    # Productos Más Rentados
    # ================================
    ws.append([])
    ws.append(["Productos Más Rentados"])
    ws[f"A{ws.max_row}"].font = title_font
    ws[f"A{ws.max_row}"].alignment = align_center

    ws.append(["Producto", "Cantidad Rentada"])
    top_rented_products = RentalOrderItem.objects.values('product_name').annotate(
        total_rented=Sum('amount')
    ).order_by('-total_rented')[:10]
    for product in top_rented_products:
        ws.append([product['product_name'], product['total_rented']])

    # Ajustar ancho de columnas automáticamente
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            # Obtener la letra de la columna
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_ordenes.xlsx"'

    # Guardar el archivo Excel en la respuesta
    wb.save(response)
    return response